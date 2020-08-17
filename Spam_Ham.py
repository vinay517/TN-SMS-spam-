import openpyxl
import csv
import re
import xlrd
from itertools import groupby
wb = openpyxl.load_workbook('SMSSpamCollection.xlsx') #loading the workbook which has the main dataset
new = openpyxl.load_workbook('new.xlsx') #a new excel sheet to enter into

sheet_Main = wb.get_sheet_by_name('Sheet1')
sheet_New = new.get_sheet_by_name('Sheet1')

#accessing each message from the dataset
for i in range(1,5575):
    message = (sheet_Main.cell(row=i+1,column=2).value)
    numbers = sum(c.isdigit() for c in message) #counts the number of NUMBERS
    words = sum(c.isalpha() for c in message) #counts the number of characters
    spaces = sum(c.isspace() for c in message) #counts the number of spaces
    currency = (message.count('$')) + (message.count('£')) #counts the currency symbols

    sheet_New.cell(row=i, column=1).value = len(message) #Gives the First Feature - Number of Characters typed in a message
    sheet_New.cell(row=i, column=2).value = currency #Gives the Second Feature - Number of currency symbols used

    #to find the third feature - number of Numeric Strings
    strrr = re.sub('\W+', ' ', message)
    regex = re.compile(r'(\d+)')
    str2 = regex.split(strrr)
    b = len(str2)
    f = sum(c.isnumeric() for c in str2)
    y = len(str2)
    sum1 = 0
    for j in range(0, y):
        if (str2[j].isnumeric()):
            sum1 = sum1 + 1
    sheet_New.cell(row=i, column=3).value = sum1 #gives the Third Feature -
    #to find the Fourth Feature - Frequency of the most Frequent word
    word_counter = {}
    message = message.lower()
    for word in message.split():
        if word in word_counter:
            word_counter[word] += 1
        else:
            word_counter[word] = 1
    highest_words = []
    highest_value = 0
    for k, v in word_counter.items():
        if v > highest_value:
            highest_words = []
            highest_words.append(k)
            highest_value = v
        elif v == highest_value:
            highest_words.append(k)
    sheet_New.cell(row=i, column=4).value = highest_value #Gives the Fourth Feature
#To add the Class of the data to our new Data Set which contains Features
for j in range(1,5575):
    sheet_New.cell(row=j, column=5).value = sheet_Main.cell(row=j + 1, column=1).value
new.save('new.xlsx')
with open('newcsv1.csv', 'w') as f:
    c = csv.writer(f)
    for j in sheet_New.rows:
       c.writerow([cell.value for cell in j])
                                                    #To write the excel sheet into a csv file
with open('newcsv1.csv') as input, open('newcsv2.csv', 'w') as output:
    non_blank = (line for line in input if line.strip())
    output.writelines(non_blank)
    input.close()
    output.close()
                                                    #generating an arff file from the csv files
file_csv = "newcsv2.csv"
file_arff = "Final_arff.arff"
myfile = csv.reader(open(file_csv, "rt"))
myARFF = open(file_arff, 'w+')
myARFF.write('@RELATION Weka\n\n')
myARFF.write("@ATTRIBUTE    charlength  REAL\n")
myARFF.write("@ATTRIBUTE    currencycount  REAL\n")
myARFF.write("@ATTRIBUTE    numstring  REAL\n")
myARFF.write("@ATTRIBUTE    freqword  REAL\n")
myARFF.write("@ATTRIBUTE    score   {ham,spam}\n")
myARFF.write('\n@data\n')
for row in myfile:
    myARFF.write(','.join(row) + '\n')
myARFF.close()

